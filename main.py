import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import threading
import time

class SistemaRegistroAlmuerzos:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Registro de Almuerzos")
        self.root.geometry("800x600")
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.archivo_excel = tk.StringVar(value='registro_bono.xlsx')
        self.hoy_str = datetime.today().strftime('%Y-%m-%d')
        self.wb = None
        self.ws = None
        
        # Estilos Excel
        self.negrita = Font(bold=True)
        self.centrado = Alignment(horizontal='center')
        self.relleno_encabezado = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        self.relleno_total = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        self.relleno_ausente = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        self.relleno_presente = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        self.borde = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Inicializar Excel
        self.inicializar_excel()
        
        # Crear interfaz
        self.crear_interfaz()
        
        # Actualizar estadísticas iniciales
        self.actualizar_estadisticas()
    
    def crear_interfaz(self):
        # Título principal
        titulo = tk.Label(self.root, text="🍽️ SISTEMA DE REGISTRO DE ALMUERZOS",
                        font=('Arial', 16, 'bold'), bg='#f0f0f0', fg='#2c3e50')
        titulo.pack(pady=10)
        
        # Fecha actual
        fecha_label = tk.Label(self.root, text=f"📅 Fecha: {self.hoy_str}",
                            font=('Arial', 12), bg='#f0f0f0', fg='#34495e')
        fecha_label.pack(pady=5)
        
        # Frame principal
        main_frame = tk.Frame(self.root, bg='#f0f0f0')
        main_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Frame de registro
        registro_frame = tk.LabelFrame(main_frame, text="📝 Registro de Acceso", 
                                    font=('Arial', 12, 'bold'), bg='#f0f0f0')
        registro_frame.pack(fill='x', pady=(0, 10))
        
        # Entrada QR
        tk.Label(registro_frame, text="Escanea el código QR o escribe manualmente:",
                font=('Arial', 10), bg='#f0f0f0').pack(pady=5)
        
        qr_frame = tk.Frame(registro_frame, bg='#f0f0f0')
        qr_frame.pack(pady=5)
        
        self.qr_entry = tk.Entry(qr_frame, font=('Arial', 12), width=40)
        self.qr_entry.pack(side='left', padx=(0, 5))
        self.qr_entry.bind('<Return>', self.procesar_qr)
        self.qr_entry.focus()
        
        registrar_btn = tk.Button(qr_frame, text="✅ Registrar", command=self.procesar_qr,
                                bg='#27ae60', fg='white', font=('Arial', 10, 'bold'))
        registrar_btn.pack(side='left')
        
        # Entrada manual
        manual_frame = tk.LabelFrame(registro_frame, text="Entrada Manual", bg='#f0f0f0')
        manual_frame.pack(fill='x', pady=10)
        
        manual_inner = tk.Frame(manual_frame, bg='#f0f0f0')
        manual_inner.pack(pady=5)
        
        tk.Label(manual_inner, text="Nombre:", bg='#f0f0f0').grid(row=0, column=0, sticky='w', padx=5)
        self.nombre_entry = tk.Entry(manual_inner, font=('Arial', 10), width=25)
        self.nombre_entry.grid(row=0, column=1, padx=5)
        
        tk.Label(manual_inner, text="Grupo:", bg='#f0f0f0').grid(row=0, column=2, sticky='w', padx=5)
        self.grupo_entry = tk.Entry(manual_inner, font=('Arial', 10), width=25)
        self.grupo_entry.grid(row=0, column=3, padx=5)
        
        manual_btn = tk.Button(manual_inner, text="📝 Registrar Manual", 
                            command=self.registrar_manual,
                            bg='#3498db', fg='white', font=('Arial', 10, 'bold'))
        manual_btn.grid(row=0, column=4, padx=10)
        
        # Frame de estadísticas y log
        stats_log_frame = tk.Frame(main_frame, bg='#f0f0f0')
        stats_log_frame.pack(fill='both', expand=True)
        
        # Estadísticas
        stats_frame = tk.LabelFrame(stats_log_frame, text="📊 Estadísticas del Día", 
                                font=('Arial', 12, 'bold'), bg='#f0f0f0')
        stats_frame.pack(side='left', fill='both', expand=True, padx=(0, 5))
        
        self.stats_text = tk.Text(stats_frame, height=8, width=35, font=('Arial', 10),
                                bg='#ecf0f1', state='disabled')
        self.stats_text.pack(padx=5, pady=5, fill='both', expand=True)
        
        # Log de actividad
        log_frame = tk.LabelFrame(stats_log_frame, text="📋 Log de Actividad", 
                                font=('Arial', 12, 'bold'), bg='#f0f0f0')
        log_frame.pack(side='right', fill='both', expand=True, padx=(5, 0))
        
        # Scrollbar para el log
        log_scroll_frame = tk.Frame(log_frame, bg='#f0f0f0')
        log_scroll_frame.pack(fill='both', expand=True, padx=5, pady=5)
        
        self.log_text = tk.Text(log_scroll_frame, height=8, width=40, font=('Arial', 9),
                            bg='#ecf0f1', state='disabled')
        log_scrollbar = tk.Scrollbar(log_scroll_frame, orient='vertical', command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        self.log_text.pack(side='left', fill='both', expand=True)
        log_scrollbar.pack(side='right', fill='y')
        
        # Frame de botones inferiores
        buttons_frame = tk.Frame(main_frame, bg='#f0f0f0')
        buttons_frame.pack(fill='x', pady=10)
        
        # Botones de acción
        tk.Button(buttons_frame, text="📊 Generar Resumen", command=self.generar_resumen,
                bg='#9b59b6', fg='white', font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        
        tk.Button(buttons_frame, text="📁 Abrir Excel", command=self.abrir_excel,
                bg='#16a085', fg='white', font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        
        tk.Button(buttons_frame, text="🔄 Actualizar Stats", command=self.actualizar_estadisticas,
                bg='#f39c12', fg='white', font=('Arial', 10, 'bold')).pack(side='left', padx=5)
        
        tk.Button(buttons_frame, text="🚪 Finalizar Día", command=self.finalizar_dia,
                bg='#e74c3c', fg='white', font=('Arial', 10, 'bold')).pack(side='right', padx=5)
        
        # Estado del archivo
        self.estado_label = tk.Label(self.root, text=f"📁 Archivo: {self.archivo_excel.get()}",
                                    font=('Arial', 9), bg='#f0f0f0', fg='#7f8c8d')
        self.estado_label.pack(pady=5)
        
        # Log inicial
        self.agregar_log("Sistema iniciado correctamente", "info")
    
    def inicializar_excel(self):
        """Inicializa el archivo Excel"""
        try:
            if not os.path.exists(self.archivo_excel.get()):
                self.wb = openpyxl.Workbook()
                self.wb.remove(self.wb.active)
                self.ws = self.wb.create_sheet('Accesos')
                self.ws.cell(1, 1).value = 'Grupo/Nombre'
                self.aplicar_estilo_celda(self.ws.cell(1, 1), 'encabezado')
                self.wb.save(self.archivo_excel.get())
            
            self.wb = openpyxl.load_workbook(self.archivo_excel.get())
            self.ws = self.wb['Accesos']
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al inicializar Excel: {e}")
    
    def agregar_log(self, mensaje, tipo="info"):
        """Agrega mensaje al log con timestamp"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        
        # Iconos según tipo
        iconos = {
            "info": "ℹ️",
            "success": "✅",
            "warning": "⚠️",
            "error": "❌"
        }
        
        icono = iconos.get(tipo, "ℹ️")
        log_mensaje = f"[{timestamp}] {icono} {mensaje}\n"
        
        self.log_text.config(state='normal')
        self.log_text.insert(tk.END, log_mensaje)
        self.log_text.config(state='disabled')
        self.log_text.see(tk.END)
    
    def normalizar(self, texto):
        """Normaliza texto para comparaciones"""
        return str(texto).lower().strip() if texto else ""
    
    def aplicar_estilo_celda(self, celda, tipo='normal'):
        """Aplica estilos a las celdas del Excel"""
        celda.border = self.borde
        if tipo == 'encabezado':
            celda.font = self.negrita
            celda.fill = self.relleno_encabezado
            celda.alignment = self.centrado
        elif tipo == 'total':
            celda.font = self.negrita
            celda.fill = self.relleno_total
            celda.alignment = self.centrado
        elif tipo == 'ausente':
            celda.fill = self.relleno_ausente
            celda.alignment = self.centrado
        elif tipo == 'presente':
            celda.fill = self.relleno_presente
            celda.alignment = self.centrado
        else:
            if celda.column > 1:
                celda.alignment = self.centrado
    
    def obtener_columna_fecha(self, fecha):
        """Obtiene o crea la columna para una fecha específica"""
        for c in range(2, self.ws.max_column + 2):
            cell_value = self.ws.cell(1, c).value
            if cell_value == fecha:
                return c
        
        nueva_col = self.ws.max_column + 1
        self.ws.cell(1, nueva_col).value = fecha
        self.aplicar_estilo_celda(self.ws.cell(1, nueva_col), 'encabezado')
        return nueva_col
    
    def encontrar_grupo(self, grupo_norm):
        """Encuentra la fila del encabezado de un grupo"""
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, 1).value
            if val and not str(val).startswith('  ') and not str(val).startswith('TOTAL'):
                if self.normalizar(val) == grupo_norm:
                    return r
        return None
    
    def encontrar_alumno_en_grupo(self, nombre_norm, fila_grupo):
        """Encuentra la fila de un alumno dentro de su grupo"""
        if not fila_grupo:
            return None
        
        r = fila_grupo + 1
        while r <= self.ws.max_row:
            val = self.ws.cell(r, 1).value
            if not val:
                break
            if str(val).startswith('TOTAL'):
                break
            if str(val).startswith('  '):
                nombre_celda = str(val)[2:].strip()
                if self.normalizar(nombre_celda) == nombre_norm:
                    return r
            r += 1
        return None
    
    def obtener_alumnos_grupo(self, fila_grupo):
        """Obtiene todas las filas de alumnos de un grupo"""
        alumnos = []
        if not fila_grupo:
            return alumnos
        
        r = fila_grupo + 1
        while r <= self.ws.max_row:
            val = self.ws.cell(r, 1).value
            if not val or str(val).startswith('TOTAL'):
                break
            if str(val).startswith('  '):
                alumnos.append(r)
            r += 1
        return alumnos
    
    def crear_grupo(self, grupo, posicion=None):
        """Crea un nuevo grupo en el Excel"""
        if posicion is None:
            posicion = self.ws.max_row + 2
        
        self.ws.insert_rows(posicion)
        self.ws.cell(posicion, 1).value = grupo
        self.aplicar_estilo_celda(self.ws.cell(posicion, 1), 'encabezado')
        return posicion
    
    def agregar_alumno_a_grupo(self, nombre, fila_grupo):
        """Agrega un nuevo alumno al grupo"""
        alumnos = self.obtener_alumnos_grupo(fila_grupo)
        if alumnos:
            posicion = alumnos[-1] + 1
        else:
            posicion = fila_grupo + 1
        
        if posicion <= self.ws.max_row:
            val = self.ws.cell(posicion, 1).value
            if val and str(val).startswith('TOTAL'):
                self.ws.insert_rows(posicion)
        else:
            self.ws.insert_rows(posicion)
        
        self.ws.cell(posicion, 1).value = f"  {nombre}"
        self.aplicar_estilo_celda(self.ws.cell(posicion, 1))
        return posicion
    
    def actualizar_total_grupo(self, grupo_norm, col_fecha):
        """Actualiza la línea de total del grupo"""
        fila_grupo = self.encontrar_grupo(grupo_norm)
        if not fila_grupo:
            return
        
        alumnos = self.obtener_alumnos_grupo(fila_grupo)
        
        presentes = 0
        for fila_alumno in alumnos:
            if self.ws.cell(fila_alumno, col_fecha).value:
                presentes += 1
        
        fila_total = None
        r = fila_grupo + 1
        while r <= self.ws.max_row + 1:
            if r > self.ws.max_row:
                self.ws.insert_rows(r)
                fila_total = r
                break
            val = self.ws.cell(r, 1).value
            if val and str(val).startswith('TOTAL'):
                fila_total = r
                break
            elif not val or (not str(val).startswith('  ')):
                self.ws.insert_rows(r)
                fila_total = r
                break
            r += 1
        
        grupo_name = self.ws.cell(fila_grupo, 1).value
        self.ws.cell(fila_total, 1).value = f"TOTAL {grupo_name}: {len(alumnos)} alumnos"
        self.ws.cell(fila_total, col_fecha).value = f"{presentes} presentes"
        self.aplicar_estilo_celda(self.ws.cell(fila_total, 1), 'total')
        self.aplicar_estilo_celda(self.ws.cell(fila_total, col_fecha), 'total')
    
    def marcar_ausencias(self, col_fecha, fecha):
        """Marca en rojo las ausencias del día"""
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        hoy_obj = datetime.today().date()
        
        if fecha_obj > hoy_obj:
            return
        
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, 1).value
            if val and str(val).startswith('  '):
                celda = self.ws.cell(r, col_fecha)
                if not celda.value:
                    self.aplicar_estilo_celda(celda, 'ausente')
                else:
                    self.aplicar_estilo_celda(celda, 'presente')
    
    def ajustar_ancho_columnas(self):
        """Ajusta el ancho de las columnas"""
        for col in self.ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0 for c in col), default=10)
            letra = col[0].column_letter
            self.ws.column_dimensions[letra].width = min(max_len + 3, 20)
    
    def registrar_acceso(self, nombre, grupo):
        """Función principal para registrar un acceso"""
        try:
            nom_norm = self.normalizar(nombre)
            grp_norm = self.normalizar(grupo)
            
            col_hoy = self.obtener_columna_fecha(self.hoy_str)
            
            fila_grupo = self.encontrar_grupo(grp_norm)
            if not fila_grupo:
                fila_grupo = self.crear_grupo(grupo)
                self.agregar_log(f"Grupo '{grupo}' creado", "info")
            
            fila_alumno = self.encontrar_alumno_en_grupo(nom_norm, fila_grupo)
            if not fila_alumno:
                fila_alumno = self.agregar_alumno_a_grupo(nombre, fila_grupo)
                self.agregar_log(f"Alumno '{nombre}' agregado al grupo", "info")
            
            if self.ws.cell(fila_alumno, col_hoy).value:
                hora_anterior = self.ws.cell(fila_alumno, col_hoy).value
                messagebox.showwarning("Ya Registrado", 
                                    f"{nombre} ya fue registrado hoy a las {hora_anterior}")
                self.agregar_log(f"DUPLICADO: {nombre} ya registrado a las {hora_anterior}", "warning")
                return False
            
            hora = datetime.now().strftime('%H:%M:%S')
            self.ws.cell(fila_alumno, col_hoy).value = hora
            self.aplicar_estilo_celda(self.ws.cell(fila_alumno, col_hoy), 'presente')
            
            self.actualizar_total_grupo(grp_norm, col_hoy)
            self.marcar_ausencias(col_hoy, self.hoy_str)
            self.ajustar_ancho_columnas()
            self.wb.save(self.archivo_excel.get())
            
            self.agregar_log(f"REGISTRADO: {nombre} ({grupo}) a las {hora}", "success")
            return True
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al registrar: {e}")
            self.agregar_log(f"ERROR: {e}", "error")
            return False
    
    def procesar_qr(self, event=None):
        """Procesa el código QR escaneado"""
        txt = self.qr_entry.get().strip()
        self.qr_entry.delete(0, tk.END)
        
        if not txt:
            return
        
        if "|" not in txt:
            messagebox.showerror("Formato Inválido", 
                            "El formato debe ser: Nombre|Grupo")
            self.agregar_log("Formato inválido en QR", "error")
            return
        
        try:
            nombre, grupo = txt.split("|", 1)
            nombre = nombre.strip()
            grupo = grupo.strip()
            
            if not nombre or not grupo:
                messagebox.showerror("Datos Incompletos", 
                                "Nombre y grupo no pueden estar vacíos")
                return
            
            if self.registrar_acceso(nombre, grupo):
                self.actualizar_estadisticas()
                # Enfocar de nuevo el campo QR
                self.qr_entry.focus()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar QR: {e}")
    
    def registrar_manual(self):
        """Registra un acceso manual"""
        nombre = self.nombre_entry.get().strip()
        grupo = self.grupo_entry.get().strip()
        
        if not nombre or not grupo:
            messagebox.showerror("Datos Incompletos", 
                            "Por favor complete nombre y grupo")
            return
        
        if self.registrar_acceso(nombre, grupo):
            self.nombre_entry.delete(0, tk.END)
            self.grupo_entry.delete(0, tk.END)
            self.actualizar_estadisticas()
            self.qr_entry.focus()
    
    def actualizar_estadisticas(self):
        """Actualiza las estadísticas mostradas"""
        try:
            col_hoy = self.obtener_columna_fecha(self.hoy_str)
            
            grupos_stats = {}
            total_presentes = 0
            total_alumnos = 0
            
            for r in range(2, self.ws.max_row + 1):
                val = self.ws.cell(r, 1).value
                if val and not str(val).startswith('  ') and not str(val).startswith('TOTAL'):
                    # Es un grupo
                    grupo = val
                    alumnos = self.obtener_alumnos_grupo(r)
                    
                    presentes = sum(1 for fila_alumno in alumnos 
                                if self.ws.cell(fila_alumno, col_hoy).value)
                    
                    grupos_stats[grupo] = {
                        'presentes': presentes,
                        'total': len(alumnos)
                    }
                    total_presentes += presentes
                    total_alumnos += len(alumnos)
            
            # Actualizar texto de estadísticas
            stats_texto = f"📊 ESTADÍSTICAS DEL DÍA\n"
            stats_texto += f"{'='*30}\n\n"
            
            for grupo, stats in grupos_stats.items():
                porcentaje = (stats['presentes'] / stats['total'] * 100) if stats['total'] > 0 else 0
                stats_texto += f"📚 {grupo}:\n"
                stats_texto += f"   Presentes: {stats['presentes']}/{stats['total']} ({porcentaje:.1f}%)\n\n"
            
            stats_texto += f"{'='*30}\n"
            stats_texto += f"🎯 TOTAL GENERAL:\n"
            stats_texto += f"   {total_presentes}/{total_alumnos} alumnos\n"
            
            if total_alumnos > 0:
                porcentaje_total = (total_presentes / total_alumnos * 100)
                stats_texto += f"   Asistencia: {porcentaje_total:.1f}%"
            
            self.stats_text.config(state='normal')
            self.stats_text.delete(1.0, tk.END)
            self.stats_text.insert(1.0, stats_texto)
            self.stats_text.config(state='disabled')
            
        except Exception as e:
            self.agregar_log(f"Error actualizando estadísticas: {e}", "error")
    
    def generar_resumen(self):
        """Genera la hoja de resumen"""
        try:
            # Usar la función original del código
            if 'Resumen' in self.wb.sheetnames:
                self.wb.remove(self.wb['Resumen'])
            
            res = self.wb.create_sheet('Resumen')
            
            fechas = []
            for c in range(2, self.ws.max_column + 1):
                fecha = self.ws.cell(1, c).value
                if fecha:
                    fechas.append(fecha)
            
            res.cell(1, 1).value = 'Grupo'
            for i, fecha in enumerate(fechas, start=2):
                res.cell(1, i).value = fecha
                self.aplicar_estilo_celda(res.cell(1, i), 'encabezado')
            res.cell(1, len(fechas) + 2).value = 'Total Alumnos'
            self.aplicar_estilo_celda(res.cell(1, 1), 'encabezado')
            self.aplicar_estilo_celda(res.cell(1, len(fechas) + 2), 'encabezado')
            
            fila_resumen = 2
            for r in range(2, self.ws.max_row + 1):
                val = self.ws.cell(r, 1).value
                if val and not str(val).startswith('  ') and not str(val).startswith('TOTAL'):
                    grupo = val
                    alumnos = self.obtener_alumnos_grupo(r)
                    
                    res.cell(fila_resumen, 1).value = grupo
                    
                    for i, fecha in enumerate(fechas, start=2):
                        col_fecha = i
                        presentes = sum(1 for fila_alumno in alumnos 
                                    if self.ws.cell(fila_alumno, col_fecha).value)
                        res.cell(fila_resumen, i).value = f"{presentes}/{len(alumnos)}"
                        self.aplicar_estilo_celda(res.cell(fila_resumen, i))
                    
                    res.cell(fila_resumen, len(fechas) + 2).value = len(alumnos)
                    self.aplicar_estilo_celda(res.cell(fila_resumen, len(fechas) + 2))
                    
                    fila_resumen += 1
            
            # Totales generales
            res.cell(fila_resumen, 1).value = 'TOTAL GENERAL'
            self.aplicar_estilo_celda(res.cell(fila_resumen, 1), 'total')
            
            for i, fecha in enumerate(fechas, start=2):
                col_fecha = i
                total_dia = 0
                for r in range(2, self.ws.max_row + 1):
                    val = self.ws.cell(r, 1).value
                    if val and str(val).startswith('  '):
                        if self.ws.cell(r, col_fecha).value:
                            total_dia += 1
                res.cell(fila_resumen, i).value = total_dia
                self.aplicar_estilo_celda(res.cell(fila_resumen, i), 'total')
            
            self.wb.save(self.archivo_excel.get())
            messagebox.showinfo("Éxito", "Resumen generado exitosamente")
            self.agregar_log("Resumen generado exitosamente", "success")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error generando resumen: {e}")
            self.agregar_log(f"Error generando resumen: {e}", "error")
    
    def abrir_excel(self):
        """Abre el archivo Excel"""
        try:
            os.startfile(self.archivo_excel.get())
            self.agregar_log("Archivo Excel abierto", "info")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {e}")
    
    def finalizar_dia(self):
        """Finaliza el día marcando ausencias y generando resumen"""
        respuesta = messagebox.askyesno("Finalizar Día", 
                                    "¿Está seguro de que desea finalizar el día?\n"
                                    "Esto marcará las ausencias y generará el resumen final.")
        if not respuesta:
            return
        
        try:
            self.agregar_log("Finalizando día...", "info")
            
            # Marcar ausencias en todas las fechas
            for c in range(2, self.ws.max_column + 1):
                fecha = self.ws.cell(1, c).value
                if fecha:
                    self.marcar_ausencias(c, fecha)
            
            self.ajustar_ancho_columnas()
            self.wb.save(self.archivo_excel.get())
            self.generar_resumen()
            self.actualizar_estadisticas()
            self.agregar_log("Día finalizado correctamente", "success")
            messagebox.showinfo("Finalizado", "El día ha sido finalizado y el resumen generado")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error finalizando el día: {e}")
            self.agregar_log(f"Error finalizando el día: {e}", "error")
    
def ordenar_alumnos_por_grupo(self):
    """Ordena alfabéticamente los alumnos dentro de cada grupo moviendo las filas en el Excel"""
    try:
        grupos = []
        # Encuentra las filas de los grupos
        for r in range(2, self.ws.max_row + 1):
            val = self.ws.cell(r, 1).value
            if val and not str(val).startswith('  ') and not str(val).startswith('TOTAL'):
                grupos.append(r)

        for fila_grupo in grupos:
            alumnos_filas = self.obtener_alumnos_grupo(fila_grupo)
            alumnos_datos = []
            for fila in alumnos_filas:
                nombre = self.ws.cell(fila, 1).value
                datos = [self.ws.cell(fila, c).value for c in range(1, self.ws.max_column + 1)]
                alumnos_datos.append((nombre, datos, fila))

            # Ordenar por nombre (ignorando los dos espacios iniciales)
            alumnos_datos.sort(key=lambda x: x[0].strip().lower() if x[0] else "")

            # Eliminar las filas de alumnos (de abajo hacia arriba para no desordenar)
            for _, _, fila in reversed(alumnos_datos):
                self.ws.delete_rows(fila)

            # Insertar los alumnos ordenados después del encabezado del grupo
            insert_pos = fila_grupo + 1
            for nombre, datos, _ in alumnos_datos:
                self.ws.insert_rows(insert_pos)
                for c, valor in enumerate(datos, start=1):
                    self.ws.cell(insert_pos, c).value = valor
                insert_pos += 1

        self.ajustar_ancho_columnas()
        self.wb.save(self.archivo_excel.get())
        self.agregar_log("Alumnos ordenados alfabéticamente en cada grupo", "success")
        messagebox.showinfo("Éxito", "Alumnos ordenados alfabéticamente en cada grupo.")
    except Exception as e:
        self.agregar_log(f"Error al ordenar alumnos: {e}", "error")
        messagebox.showerror("Error", f"Error al ordenar alumnos: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = SistemaRegistroAlmuerzos(root)
    root.mainloop()